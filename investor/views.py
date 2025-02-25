from openpyxl import load_workbook
from django.http import JsonResponse
from django.core.paginator import Paginator
from resources.utils import get_supabase_client
from rest_framework.views import APIView

class InvestorResourceList(APIView):
    def get(self, request, *args, **kwargs):
        supabase = get_supabase_client()

        # Fetch query parameters for pagination
        page = int(request.GET.get("page", 1))  # Default to page 1 if not provided
        per_page = int(request.GET.get("per_page", 10))  # Default to 10 items per page

        # Get data from Supabase
        response = supabase.table("Investor").select("*").execute()

        # Full dataset from Supabase
        data = response.data

        # Use Django Paginator
        paginator = Paginator(data, per_page)
        try:
            paginated_data = paginator.page(page)
        except:
            return JsonResponse({"error": "Invalid page number"}, status=404)

        # Return paginated response
        return JsonResponse({
            "data": list(paginated_data),
            "total_items": paginator.count,
            "total_pages": paginator.num_pages,
            "current_page": page
        })



def CreateInvestorResources(request):

    # Load the workbook
    file_path = "Book2.xlsx"  # Replace with your file path
    workbook = load_workbook(file_path)

    # Access a sheet by name
    sheet = workbook["Sheet1"]
    
    # Iterate through rows
    rows = []
    for row in sheet.iter_rows(values_only=True): # Each row as a tuple
        rows.append(row)
    
    i = 0
    j = len(rows)
    articles = []
    videos = []
    supabase = get_supabase_client()
    while j:
        if rows[i][2]:
            articles.append(rows[i][2])
        if rows[i][3]:
            videos.append(rows[i][3])
        prompt = rows[i][1]
        i += 1
        while rows[i][0] == None:
            if rows[i][2]:
                articles.append(rows[i][2])
            if rows[i][3]:
                videos.append(rows[i][3])
            i += 1
        print(prompt, articles, videos)
        supabase.table("Investor").insert({
            "prompt": prompt,
            "articles": articles,
            "videos": videos,
        }).execute()
        articles = []
        videos = []
        j -= 1