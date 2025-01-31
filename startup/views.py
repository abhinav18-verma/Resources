from rest_framework.views import APIView
from rest_framework.response import Response
from .models import Resource
from .serializers import ResourceSerializer
from openpyxl import load_workbook
from rest_framework.pagination import LimitOffsetPagination
from supabase import create_client
from django.conf import settings
from django.http import JsonResponse
from django.core.paginator import Paginator

class ResourceListView(APIView):
    def get(self, request):
        resources = Resource.objects.all()
    
        if resources:
            paginator = LimitOffsetPagination()
            paginated_list = paginator.paginate_queryset(
                resources, request)
            data = ResourceSerializer(
                paginated_list, many=True).data
        else:
            paginator, data = None, []
    
        return Response(
            {
                "count": paginator.count,
                "next": paginator.get_next_link(),
                "previous": paginator.get_previous_link(),
                "result": data,
                "message": "Startup resources fetched successfully."
            },
            status=200
        )



def CreateStartupResources(request):

    # Load the workbook
    file_path = "Book1.xlsx"  # Replace with your file path
    workbook = load_workbook(file_path)

    # Access a sheet by name
    sheet = workbook["Sheet1"]
    
    # Iterate through rows
    rows = []
    for row in sheet.iter_rows(values_only=True): # Each row as a tuple
        rows.append(row)
    
    i = 0
    j = len(rows)
    articles = []
    videos = []
    while j:
        if rows[i][2]:
            articles.append(rows[i][2])
        if rows[i][3]:
            videos.append(rows[i][3])
        prompt = rows[i][1]
        i += 1
        while rows[i][0] == None:
            if rows[i][2]:
                articles.append(rows[i][2])
            if rows[i][3]:
                videos.append(rows[i][3])
            i += 1
        print(prompt, articles, videos)
        Resource.objects.create(
            prompt = prompt,
            articles = articles,
            videos = videos
        )
        articles = []
        videos = []
        j -= 1

def get_supabase_client():
    url = settings.SUPABASE_URL
    key = settings.SUPABASE_KEY
    return create_client(url, key)

def add_data(request):
    supabase = get_supabase_client()
    data = {
        "name": "Sample Item 2"
    }
    response = supabase.table("Person").insert(data).execute()
    return JsonResponse(response.data, safe=False)

def fetch_paginated_data(request):
    supabase = get_supabase_client()

    # Fetch query parameters for pagination
    page = int(request.GET.get("page", 1))  # Default to page 1 if not provided
    per_page = int(request.GET.get("per_page", 10))  # Default to 10 items per page

    # Get data from Supabase
    response = supabase.table("Person").select("*").execute()

    # Full dataset from Supabase
    data = response.data

    # Use Django Paginator
    paginator = Paginator(data, per_page)
    try:
        paginated_data = paginator.page(page)
    except:
        return JsonResponse({"error": "Invalid page number"}, status=404)

    # Return paginated response
    return JsonResponse({
        "data": list(paginated_data),
        "total_items": paginator.count,
        "total_pages": paginator.num_pages,
        "current_page": page
    })
