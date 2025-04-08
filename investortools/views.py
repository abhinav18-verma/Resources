from django.shortcuts import render
from openpyxl import load_workbook
from resources.utils import get_supabase_client
from rest_framework.views import APIView
from django.core.paginator import Paginator
from django.http import JsonResponse
# Create your views here.

class InvestorToolsList(APIView):
    def get(self, request, *args, **kwargs):
        supabase = get_supabase_client()

        # Fetch query parameters for pagination
        page = int(request.GET.get("page", 1))  # Default to page 1 if not provided
        per_page = int(request.GET.get("per_page", 10))  # Default to 10 items per page

        # Get data from Supabase
        response = supabase.table("InvestorTools").select("*").execute()

        # Full dataset from Supabase
        data = response.data

        # Use Django Paginator
        paginator = Paginator(data, per_page)
        try:
            paginated_data = paginator.page(page)
        except:
            return JsonResponse({"error": "Invalid page number"}, status=404)

        # Return paginated response
        return JsonResponse({
            "data": list(paginated_data),
            "total_items": paginator.count,
            "total_pages": paginator.num_pages,
            "current_page": page
        })
    
def CreateInvestorTools(request):

    file_path = "Book6.xlsx"  # Replace with your file path
    workbook = load_workbook(file_path)

    # Access a sheet by name
    sheet = workbook["Sheet1"]
    
    supabase = get_supabase_client()

    # Iterate through rows
    for row in sheet.iter_rows(values_only=True): # Each row as a tuple
        supabase.table("InvestorTools").insert({
            "Serial Number": row[0],
            "Title": row[1],
            "Description": row[2],
            "Link": row[3],
        }).execute()